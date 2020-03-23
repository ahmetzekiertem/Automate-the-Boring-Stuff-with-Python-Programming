import openpyxl


wb = openpyxl.load_workbook('example.xlsx')

type(wb)

wb.get_sheet_names()
['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb.get_sheet_by_name('Sheet3')

sheet


sheet.title
#'Sheet3'
anotherSheet = wb.active
anotherSheet

sheet['A1']

sheet['A1'].value
#datetime.datetime(2015, 4, 5, 13, 34, 2)
c = sheet['B1']
c.value
str(sheet['C1'].value)
#Apples'
print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
'Row 1, Column B is Apples'
print('Cell ' + c.coordinate + ' is ' + c.value)
'Cell B1 is Apples'
sheet['C1'].value
73


sheet.cell(row=1, column=2)

sheet.cell(row=1, column=2).value

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

'''
1 Apples
3 Pears
5 Apples
7 Strawberries
'''

'''You can determine the size of the sheet with the Worksheet object’s max_row and
 max_column member variables. Enter the following into the interactive shell:'''


import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
sheet.max_row

sheet.max_column


 '''To convert from numbers to letters, call the openpyxl.cell.get_column_letter() function.
  Enter the following into the interactive shell:'''


import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string

get_column_letter(1)

get_column_letter(2)

get_column_letter(27)

get_column_letter(900)

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
get_column_letter(sheet.max_column)


column_index_from_string('A')

column_index_from_string('AA')


wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

sheet["A1"].value

sheet["A1"].value == None

sheet["A1"].value = 42

sheet["A3"].value = 'Hello'

os.chdir("/Users/mac/Desktop")
wb.save('excel.xlsx')


tuple(sheet['A1':'C3'])
  ''' ((<Cell Sheet1.A1>, <Cell Sheet1.B1>, <Cell Sheet1.C1>), (<Cell Sheet1.A2>,
   <Cell Sheet1.B2>, <Cell Sheet1.C2>), (<Cell Sheet1.A3>, <Cell Sheet1.B3>,
   <Cell Sheet1.C3>))'''

  for rowOfCellObjects in sheet['A1':'C3']:
      for cellObj in rowOfCellObjects:
          print(cellObj.coordinate, cellObj.value)
           print('--- END OF ROW ---')

'''
A1 2015-04-05 13:34:02
   B1 Apples
   C1 73
   --- END OF ROW ---
   A2 2015-04-05 03:41:23
   B2 Cherries
   C2 85
   --- END OF ROW ---
   A3 2015-04-06 12:46:51
   B3 Pears
   C3 14
   --- END OF ROW ---
   '''

sheet = wb.active
sheet.columns[1]
'''(<Cell Sheet1.B1>, <Cell Sheet1.B2>, <Cell Sheet1.B3>, <Cell Sheet1.B4>,
<Cell Sheet1.B5>, <Cell Sheet1.B6>, <Cell Sheet1.B7>)'''

for cellObj in sheet.columns[1]:
    print(cellObj.value)

'''
Apples
Cherries
Pears
Oranges
Apples
Bananas
Strawberries'''

 #! python3
   # readCensusExcel.py - Tabulates population and number of census tracts for
   # each county.


import openpyxl, pprint
print('Opening workbook...')

wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.get_sheet_by_name('Population by Census Tract')

countyData = {}

   # TODO: Fill in countyData with each county's population and tracts.
print('Reading rows...')
